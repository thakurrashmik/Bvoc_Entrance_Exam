from __future__ import annotations

import json
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_FILE = DATA_DIR / "exam_portal.xlsx"
RESULTS_FILE = BASE_DIR / "Exam_result.xlsx"

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-this-secret-key")

SHEETS = {
    "Users": ["id", "name", "email", "password_hash", "role"],
    "Questions": [
        "id",
        "question",
        "options_json",
        "correct_answer_json",
        "question_type",
        "subject",
        "topic",
        "difficulty",
    ],
    "Exams": [
        "id",
        "title",
        "duration_minutes",
        "total_marks",
        "negative_marking",
        "published",
        "question_ids_json",
    ],
}

RESULT_HEADERS = [
    "id",
    "student_id",
    "student_name",
    "exam_id",
    "exam_title",
    "score",
    "correct_count",
    "incorrect_count",
    "submitted_at",
    "answers_json",
]

DEFAULT_ADMIN = {
    "id": "admin-1",
    "name": "Admin User",
    "email": "admin@example.com",
    "password": "admin123",
    "role": "admin",
}

DEFAULT_STUDENT = {
    "id": "student-1",
    "name": "Student User",
    "email": "student@example.com",
    "password": "student123",
    "role": "student",
}

DEFAULT_QUESTIONS = [
    {
        "id": "q1",
        "question": "What is 2 + 2?",
        "options": ["3", "4", "5", "6"],
        "correct_answer": ["4"],
        "question_type": "mcq",
        "subject": "Mathematics",
        "topic": "Arithmetic",
        "difficulty": "Easy",
    },
    {
        "id": "q2",
        "question": "Which of the following are prime numbers?",
        "options": ["2", "3", "4", "9"],
        "correct_answer": ["2", "3"],
        "question_type": "multi",
        "subject": "Mathematics",
        "topic": "Number Theory",
        "difficulty": "Medium",
    },
    {
        "id": "q3",
        "question": "Flask is a Python web framework.",
        "options": ["True", "False"],
        "correct_answer": ["True"],
        "question_type": "true_false",
        "subject": "Programming",
        "topic": "Python",
        "difficulty": "Easy",
    },
]

DEFAULT_EXAMS = [
    {
        "id": "exam-1",
        "title": "Starter Aptitude Test",
        "duration_minutes": 10,
        "total_marks": 3,
        "negative_marking": 0,
        "published": 1,
        "question_ids": ["q1", "q2", "q3"],
    }
]


def ensure_storage() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DATA_FILE.exists():
        workbook = load_workbook(DATA_FILE)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    for sheet_name, headers in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(headers)
        else:
            worksheet = workbook[sheet_name]
            if worksheet.max_row == 0:
                worksheet.append(headers)
            elif [cell.value for cell in worksheet[1]] != headers:
                existing_headers = [cell.value for cell in worksheet[1]]
                if existing_headers != headers:
                    worksheet.delete_rows(1, 1)
                    worksheet.insert_rows(1)
                    worksheet.append(headers)

    workbook.save(DATA_FILE)


def workbook() -> Any:
    ensure_storage()
    return load_workbook(DATA_FILE)


def results_workbook() -> Any:
    should_save = False
    if RESULTS_FILE.exists():
        workbook = load_workbook(RESULTS_FILE)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
        should_save = True

    if "Results" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("Results")
        worksheet.append(RESULT_HEADERS)
        should_save = True
    else:
        worksheet = workbook["Results"]
        if worksheet.max_row == 0:
            worksheet.append(RESULT_HEADERS)
            should_save = True
        elif [cell.value for cell in worksheet[1]] != RESULT_HEADERS:
            worksheet.delete_rows(1, 1)
            worksheet.insert_rows(1)
            worksheet.append(RESULT_HEADERS)
            should_save = True

    if should_save:
        workbook.save(RESULTS_FILE)
    return load_workbook(RESULTS_FILE)


def sheet_rows(sheet_name: str) -> list[dict[str, Any]]:
    wb = workbook()
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        rows.append({header: row[index] for index, header in enumerate(headers)})
    return rows


def result_rows() -> list[dict[str, Any]]:
    wb = results_workbook()
    ws = wb["Results"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        rows.append({header: row[index] for index, header in enumerate(headers)})
    return rows


def save_sheet_row(sheet_name: str, values: dict[str, Any]) -> None:
    wb = workbook()
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header) for header in headers])
    wb.save(DATA_FILE)


def save_result_row(values: dict[str, Any]) -> None:
    wb = results_workbook()
    ws = wb["Results"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header) for header in headers])
    wb.save(RESULTS_FILE)


def update_row(sheet_name: str, row_id: str, values: dict[str, Any]) -> None:
    wb = workbook()
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    id_index = headers.index("id") + 1
    for row_number in range(2, ws.max_row + 1):
        if ws.cell(row=row_number, column=id_index).value == row_id:
            for column_number, header in enumerate(headers, start=1):
                if header in values:
                    ws.cell(row=row_number, column=column_number, value=values[header])
            wb.save(DATA_FILE)
            return


def get_user_by_email(email: str) -> dict[str, Any] | None:
    for user in sheet_rows("Users"):
        if (user.get("email") or "").strip().lower() == email.strip().lower():
            return user
    return None


def get_user_by_id(user_id: str) -> dict[str, Any] | None:
    for user in sheet_rows("Users"):
        if user.get("id") == user_id:
            return user
    return None


def get_question_by_id(question_id: str) -> dict[str, Any] | None:
    for question in sheet_rows("Questions"):
        if question.get("id") == question_id:
            return question
    return None


def get_exam_by_id(exam_id: str) -> dict[str, Any] | None:
    for exam in sheet_rows("Exams"):
        if exam.get("id") == exam_id:
            return exam
    return None


def json_load(value: Any, fallback: Any) -> Any:
    if value in (None, ""):
        return fallback
    if isinstance(value, (list, dict)):
        return value
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return fallback


def current_user() -> dict[str, Any] | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_user_by_id(user_id)


def login_required(view):
    def wrapped(*args: Any, **kwargs: Any):
        if not current_user():
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    wrapped.__name__ = view.__name__
    return wrapped


def role_required(role: str):
    def decorator(view):
        def wrapped(*args: Any, **kwargs: Any):
            user = current_user()
            if not user or user.get("role") != role:
                flash("You do not have permission to access that page.", "danger")
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)

        wrapped.__name__ = view.__name__
        return wrapped

    return decorator


def seed_defaults() -> None:
    ensure_storage()
    results_workbook()
    if not sheet_rows("Users"):
        save_sheet_row(
            "Users",
            {
                "id": DEFAULT_ADMIN["id"],
                "name": DEFAULT_ADMIN["name"],
                "email": DEFAULT_ADMIN["email"],
                "password_hash": generate_password_hash(DEFAULT_ADMIN["password"]),
                "role": DEFAULT_ADMIN["role"],
            },
        )
        save_sheet_row(
            "Users",
            {
                "id": DEFAULT_STUDENT["id"],
                "name": DEFAULT_STUDENT["name"],
                "email": DEFAULT_STUDENT["email"],
                "password_hash": generate_password_hash(DEFAULT_STUDENT["password"]),
                "role": DEFAULT_STUDENT["role"],
            },
        )
    if not sheet_rows("Questions"):
        for question in DEFAULT_QUESTIONS:
            save_sheet_row(
                "Questions",
                {
                    "id": question["id"],
                    "question": question["question"],
                    "options_json": json.dumps(question["options"]),
                    "correct_answer_json": json.dumps(question["correct_answer"]),
                    "question_type": question["question_type"],
                    "subject": question["subject"],
                    "topic": question["topic"],
                    "difficulty": question["difficulty"],
                },
            )
    if not sheet_rows("Exams"):
        for exam in DEFAULT_EXAMS:
            save_sheet_row(
                "Exams",
                {
                    "id": exam["id"],
                    "title": exam["title"],
                    "duration_minutes": exam["duration_minutes"],
                    "total_marks": exam["total_marks"],
                    "negative_marking": exam["negative_marking"],
                    "published": exam["published"],
                    "question_ids_json": json.dumps(exam["question_ids"]),
                },
            )


def compute_exam_score(exam: dict[str, Any], submitted_answers: dict[str, Any]) -> dict[str, Any]:
    question_ids = json_load(exam.get("question_ids_json"), [])
    negative_marking = float(exam.get("negative_marking") or 0)
    correct_count = 0
    incorrect_count = 0
    score = 0.0

    for question_id in question_ids:
        question = get_question_by_id(question_id)
        if not question:
            continue
        expected = sorted(json_load(question.get("correct_answer_json"), []))
        response = submitted_answers.get(question_id)
        if response is None:
            incorrect_count += 1
            continue
        if isinstance(response, list):
            normalized = sorted([str(item) for item in response if str(item).strip()])
        else:
            normalized = [str(response)] if str(response).strip() else []
        if normalized == expected:
            correct_count += 1
            score += 1
        else:
            incorrect_count += 1
            score -= negative_marking

    return {
        "score": max(score, 0),
        "correct_count": correct_count,
        "incorrect_count": incorrect_count,
    }


@app.context_processor
def inject_user() -> dict[str, Any]:
    return {"logged_in_user": current_user()}


@app.route("/")
def index():
    user = current_user()
    if user:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "student")

        if not name or not email or not password:
            flash("All registration fields are required.", "danger")
            return redirect(url_for("register"))
        if get_user_by_email(email):
            flash("An account with that email already exists.", "warning")
            return redirect(url_for("register"))

        save_sheet_row(
            "Users",
            {
                "id": str(uuid4()),
                "name": name,
                "email": email,
                "password_hash": generate_password_hash(password),
                "role": role if role in {"admin", "student"} else "student",
            },
        )
        flash("Account created. Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        user = get_user_by_email(email)
        if not user or not check_password_hash(user.get("password_hash", ""), password):
            flash("Invalid email or password.", "danger")
            return redirect(url_for("login"))
        session["user_id"] = user["id"]
        flash(f"Welcome back, {user['name']}.", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    exams = sheet_rows("Exams")
    questions = sheet_rows("Questions")
    if user and user.get("role") == "admin":
        all_results = result_rows()
        total_scores = [float(result.get("score") or 0) for result in all_results]
        average_score = round(sum(total_scores) / len(total_scores), 2) if total_scores else 0
        return render_template(
            "dashboard.html",
            exams=exams,
            questions=questions,
            results=all_results,
            average_score=average_score,
            published_count=len([exam for exam in exams if int(exam.get("published") or 0) == 1]),
        )
    student_results = [result for result in result_rows() if result.get("student_id") == user.get("id")]
    return render_template("dashboard.html", exams=[exam for exam in exams if int(exam.get("published") or 0) == 1], student_results=student_results)


@app.route("/admin/questions", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_questions():
    if request.method == "POST":
        question_text = request.form.get("question", "").strip()
        question_type = request.form.get("question_type", "mcq")
        subject = request.form.get("subject", "").strip()
        topic = request.form.get("topic", "").strip()
        difficulty = request.form.get("difficulty", "Easy")
        options = [value.strip() for value in request.form.getlist("options") if value.strip()]
        correct_answers = [value.strip() for value in request.form.getlist("correct_answer") if value.strip()]

        if not question_text or not options or not correct_answers:
            flash("Question text, options, and at least one correct answer are required.", "danger")
            return redirect(url_for("admin_questions"))

        save_sheet_row(
            "Questions",
            {
                "id": str(uuid4()),
                "question": question_text,
                "options_json": json.dumps(options),
                "correct_answer_json": json.dumps(correct_answers),
                "question_type": question_type,
                "subject": subject,
                "topic": topic,
                "difficulty": difficulty,
            },
        )
        flash("Question added.", "success")
        return redirect(url_for("admin_questions"))

    return render_template("questions.html", questions=sheet_rows("Questions"))


@app.route("/admin/questions/<question_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def delete_question(question_id: str):
    wb = workbook()
    ws = wb["Questions"]
    headers = [cell.value for cell in ws[1]]
    id_index = headers.index("id") + 1
    for row_number in range(2, ws.max_row + 1):
        if ws.cell(row=row_number, column=id_index).value == question_id:
            ws.delete_rows(row_number, 1)
            wb.save(DATA_FILE)
            flash("Question deleted.", "info")
            break
    return redirect(url_for("admin_questions"))


@app.route("/admin/exams", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_exams():
    questions = sheet_rows("Questions")
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        duration_minutes = int(request.form.get("duration_minutes", 10))
        total_marks = float(request.form.get("total_marks", 0) or 0)
        negative_marking = float(request.form.get("negative_marking", 0) or 0)
        published = 1 if request.form.get("published") == "on" else 0
        selected_questions = request.form.getlist("question_ids")

        if not title or not selected_questions:
            flash("Exam title and at least one question are required.", "danger")
            return redirect(url_for("admin_exams"))

        save_sheet_row(
            "Exams",
            {
                "id": str(uuid4()),
                "title": title,
                "duration_minutes": duration_minutes,
                "total_marks": total_marks,
                "negative_marking": negative_marking,
                "published": published,
                "question_ids_json": json.dumps(selected_questions),
            },
        )
        flash("Exam created.", "success")
        return redirect(url_for("admin_exams"))

    return render_template("exams.html", exams=sheet_rows("Exams"), questions=questions)


@app.route("/admin/exams/<exam_id>/toggle", methods=["POST"])
@login_required
@role_required("admin")
def toggle_exam(exam_id: str):
    exam = get_exam_by_id(exam_id)
    if exam:
        update_row("Exams", exam_id, {"published": 0 if int(exam.get("published") or 0) == 1 else 1})
        flash("Exam visibility updated.", "success")
    return redirect(url_for("admin_exams"))


@app.route("/exam/<exam_id>")
@login_required
def take_exam(exam_id: str):
    exam = get_exam_by_id(exam_id)
    if not exam or int(exam.get("published") or 0) != 1:
        flash("That exam is not available.", "warning")
        return redirect(url_for("dashboard"))

    question_ids = json_load(exam.get("question_ids_json"), [])
    questions = []
    for question_id in question_ids:
        question = get_question_by_id(question_id)
        if question:
            question["options"] = json_load(question.get("options_json"), [])
            questions.append(question)

    end_time = datetime.utcnow() + timedelta(minutes=int(exam.get("duration_minutes") or 10))
    return render_template("take_exam.html", exam=exam, questions=questions, end_time=end_time.isoformat() + "Z")


@app.route("/exam/<exam_id>/submit", methods=["POST"])
@login_required
def submit_exam(exam_id: str):
    exam = get_exam_by_id(exam_id)
    if not exam:
        flash("Exam not found.", "danger")
        return redirect(url_for("dashboard"))

    answers = {}
    for key, value in request.form.items():
        if key.startswith("question_"):
            question_id = key.replace("question_", "", 1)
            answers[question_id] = request.form.getlist(key) if len(request.form.getlist(key)) > 1 else value

    normalized_answers: dict[str, Any] = {}
    for key in request.form.keys():
        if key.startswith("question_"):
            question_id = key.replace("question_", "", 1)
            values = request.form.getlist(key)
            normalized_answers[question_id] = values if len(values) > 1 else (values[0] if values else "")

    result = compute_exam_score(exam, normalized_answers)
    user = current_user()
    result_id = str(uuid4())
    save_result_row(
        {
            "id": result_id,
            "student_id": user.get("id") if user else "",
            "student_name": user.get("name") if user else "",
            "exam_id": exam_id,
            "exam_title": exam.get("title"),
            "score": result["score"],
            "correct_count": result["correct_count"],
            "incorrect_count": result["incorrect_count"],
            "submitted_at": datetime.utcnow().isoformat(),
            "answers_json": json.dumps(normalized_answers),
        }
    )
    return redirect(url_for("view_result", result_id=result_id))


@app.route("/results/<result_id>")
@login_required
def view_result(result_id: str):
    result = next((item for item in result_rows() if item.get("id") == result_id), None)
    if not result:
        flash("Result not found.", "danger")
        return redirect(url_for("dashboard"))
    answers = json_load(result.get("answers_json"), {})
    exam = get_exam_by_id(result.get("exam_id"))
    return render_template("results.html", result=result, answers=answers, exam=exam)


@app.route("/api/exams/<exam_id>")
@login_required
def api_exam(exam_id: str):
    exam = get_exam_by_id(exam_id)
    if not exam:
        return {"error": "Exam not found"}, 404
    return {
        "id": exam.get("id"),
        "title": exam.get("title"),
        "duration_minutes": exam.get("duration_minutes"),
        "published": exam.get("published"),
        "question_ids": json_load(exam.get("question_ids_json"), []),
    }


seed_defaults()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "10000"))
    app.run(host="0.0.0.0", port=port, debug=False)
